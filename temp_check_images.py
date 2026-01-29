import openpyxl
from openpyxl.drawing.image import Image as XlImage
import os

# 讀取最新的Excel
reports_dir = 'reports'
files = [f for f in os.listdir(reports_dir) if 'ApplicationsOfDerivatives_14B_Ab3' in f and f.endswith('.xlsx')]
files.sort(reverse=True)
latest_file = os.path.join(reports_dir, files[0])

# 用 openpyxl 打開查看圖片
wb = openpyxl.load_workbook(latest_file)
ws = wb.active

print(f'檔案: {files[0]}')
print(f'工作表: {ws.title}')
print(f'行數: {ws.max_row}, 列數: {ws.max_column}')

# 檢查是否有圖片
print(f'\n圖片物件數: {len(ws._images)}')
if ws._images:
    for i, img in enumerate(ws._images, 1):
        print(f'  圖片 {i}: {img}')
        # 嘗試找到對應的單元格
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and 'latex' in str(cell.value).lower():
                    print(f'    可能關聯到: {cell.coordinate}')

# 列出前 5 行的欄位名稱和內容
print('\n前 5 列的標題和數據:')
for col_idx in range(1, min(13, ws.max_column + 1)):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    header = ws[f'{col_letter}1'].value
    print(f'  {col_letter}: {header}')
